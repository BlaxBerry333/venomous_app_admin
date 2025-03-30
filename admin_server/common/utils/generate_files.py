import csv
from datetime import datetime
import io

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


__all__ = [
    "generate_csv_file",
    "generate_excel_file",
]


def __fill_empty_rows(data, min_rows=10):
    """
    自动补充数据空行到指定数量的函数

    Parameters:
        data (list): 需要补充空行的数据列表
        min_rows (int): 补充空行的目标数量，默认为至少 10 行

    Example:
        data = [['John', 'Doe'], ['Jane', 'Doe']]
        fill_empty_rows(data, num_rows=5)
        print(data)  # 输出 [['John', 'Doe'], ['Jane', 'Doe'], ['', ''], ['', ''], ['', ''], ['', '']]
    """

    while len(data) < min_rows:
        data.append([""] * len(data[0]))


def generate_csv_file(headers, data):
    """
    生成一个 CSV 文件的内存缓冲区

    Parameters:
        headers (list): CSV 文件的表头
        data (list): CSV 文件的数据

    Returns:
        io.BytesIO: CSV 文件的内存缓冲区

    Example:
        buffer = generate_csv_file(
            headers=['id', 'name'],
            data=[[1, 'John'], [2, 'Jane']],
        )
        file_data = buffer.getvalue()
    """

    # 数据自动补充空行到 10 行
    __fill_empty_rows(data, min_rows=10)

    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # 写入 CSV 的表头
    writer.writerow(headers)

    # 写入 CSV 的数据
    for row in data:
        writer.writerow(row)

    # 确保指针回到文件开头位置
    buffer.seek(0)

    return buffer


def generate_excel_file(headers, data, sheet_name):
    """
    生成一个 Excel 文件的内存缓冲区

    Parameters:
        headers (list): Excel 文件的表头
        data (list): Excel 文件的数据

    Returns:
        io.BytesIO: Excel 文件的内存缓冲区

    Example:
        buffer = download_excel_file(
            headers=['id', 'name'],
            data=[[1, 'John'], [2, 'Jane']],
        )
        file_data = buffer.getvalue()
    """

    # 格式化数据，特别处理日期时间和 NaT
    formatted_data = []
    for row in data:
        formatted_row = []
        for value in row:
            if isinstance(value, datetime):
                formatted_row.append(
                    value.strftime("%Y-%m-%d %H:%M:%S") if value else ""
                )
            elif value is None or str(value).lower() == "nat":
                formatted_row.append("")
            else:
                formatted_row.append(str(value))
        formatted_data.append(formatted_row)

    # 数据自动补充空行到 10 行
    __fill_empty_rows(formatted_data, min_rows=10)

    df = pd.DataFrame(formatted_data, columns=headers)
    df = df.astype(str)

    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
        )
        # 获取当前的 Excel 工作表
        worksheet = writer.sheets[sheet_name]

        # 设置单元格宽度
        for row in worksheet.rows:
            worksheet.row_dimensions[row[0].row].height = 20

        # 设置单元格宽度
        for col in worksheet.columns:
            column_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[column_letter].width = 40

        # 设置单元格样式
        for row in worksheet.iter_rows():
            for cell in row:
                # 对齐
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                # 粗体
                cell.font = Font(
                    name="Arial",
                    size=12,
                    bold=True if cell.column == 1 or cell.row == 1 else False,
                )

    # 确保指针回到文件开头位置
    buffer.seek(0)

    return buffer
